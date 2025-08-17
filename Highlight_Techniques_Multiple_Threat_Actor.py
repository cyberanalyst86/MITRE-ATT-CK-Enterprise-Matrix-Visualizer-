from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from collections import defaultdict
import pandas as pd
from openpyxl.styles import Border, Side, Alignment
from collections import Counter
from print_frequency_count import *
from get_technique_information import *

def risk_color(frequency, min_freq, max_freq):
    if max_freq == min_freq:
        return "FFA500"  # fallback to orange

    scale = (frequency - min_freq) / (max_freq - min_freq)

    if scale <= 0.5:
        # Yellow to Orange
        ratio = scale / 0.5
        r = 255
        g = int(255 - ratio * (255 - 165))  # 255 → 165
        b = 0
    else:
        # Orange to Red
        ratio = (scale - 0.5) / 0.5
        r = 255
        g = int(165 - ratio * 165)  # 165 → 0
        b = 0

    return f"{r:02X}{g:02X}{b:02X}"


def Highlight_Techniques_Multiple_Threat_Actor(df_mitre):

    # Create a thin border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    input_file = r"mitre_attack_matrix_enterprise_formatted.xlsx"
    output_file = r"mitre_attack_matrix_enterprise_highlighted.xlsx"

    highlight_ids = df_mitre["mitre"].tolist()

    # Count frequency of each ID in the list
    highlight_freqs = Counter(highlight_ids)

    print_frequency_count(highlight_freqs)

    get_technique_information(highlight_freqs)

    min_freq = min(highlight_freqs.values())
    max_freq = max(highlight_freqs.values())

    wb = load_workbook(input_file)
    ws = wb.active

    # Set column widths
    for col in ws.columns:
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = 20

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.hyperlink:
                link = cell.hyperlink.target.rstrip("/")
                for tid in highlight_freqs:
                    if link.endswith(tid):
                        color_hex = risk_color(highlight_freqs[tid], min_freq, max_freq)
                        fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
                        cell.fill = fill
                        cell.border = thin_border
                        cell.alignment = Alignment(
                            wrap_text=True,
                            horizontal='center',
                            vertical='center'
                        )
                        break

    # Define gray fill for header rows
    gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    # Apply gray fill to first two rows
    for row in [1, 2]:
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.fill = gray_fill

    wb.save(output_file)

    print(f"Highlighted matrix saved to:\n{output_file}")

