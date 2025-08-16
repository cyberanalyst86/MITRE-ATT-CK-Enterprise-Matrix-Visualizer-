import openpyxl
from openpyxl.styles import Font
import re
from openpyxl.styles import Border, Side, Alignment

def format_mitre_attack_enterprise_matrix():

    input_path = r"mitre_attack_matrix_enterprise.xlsx"
    output_path = r"mitre_attack_matrix_enterprise_formatted.xlsx"

    TACTIC_LINKS = {
        "Reconnaissance": "https://attack.mitre.org/tactics/TA0043/",
        "Resource Development": "https://attack.mitre.org/tactics/TA0042/",
        "Initial Access": "https://attack.mitre.org/tactics/TA0001/",
        "Execution": "https://attack.mitre.org/tactics/TA0002/",
        "Persistence": "https://attack.mitre.org/tactics/TA0003/",
        "Privilege Escalation": "https://attack.mitre.org/tactics/TA0004/",
        "Defense Evasion": "https://attack.mitre.org/tactics/TA0005/",
        "Credential Access": "https://attack.mitre.org/tactics/TA0006/",
        "Discovery": "https://attack.mitre.org/tactics/TA0007/",
        "Lateral Movement": "https://attack.mitre.org/tactics/TA0008/",
        "Collection": "https://attack.mitre.org/tactics/TA0009/",
        "Command and Control": "https://attack.mitre.org/tactics/TA0011/",
        "Exfiltration": "https://attack.mitre.org/tactics/TA0010/",
        "Impact": "https://attack.mitre.org/tactics/TA0040/"
    }

    # Create a thin border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    wb_in = openpyxl.load_workbook(input_path)
    ws_in = wb_in.active

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active

    for col_idx in range(1, ws_in.max_column + 1):
        tactic_name = ws_in.cell(row=1, column=col_idx).value

        count_text = ws_in.cell(row=2, column=col_idx).value

        out_cell = ws_out.cell(row=1, column=col_idx)

        out_cell.value = tactic_name

        if tactic_name in TACTIC_LINKS:

            out_cell.hyperlink = TACTIC_LINKS[tactic_name]

            url = out_cell.hyperlink.target
            tactic_id = url.split("/")[-2]

            out_cell.value = str(tactic_name) + "\n" + str(tactic_id)


            out_cell.style = "Hyperlink"
            out_cell.font = Font(bold=True)

            out_cell.border = thin_border
            out_cell.alignment = Alignment(
                wrap_text=True,
                horizontal='center',
                vertical='center'
            )



        if count_text:

            ws_out.cell(row=2, column=col_idx).value = str(count_text).strip()

            out_cell = ws_out.cell(row=2, column=col_idx)

            out_cell.border = thin_border
            out_cell.alignment = Alignment(
                wrap_text=True,
                horizontal='center',
                vertical='center'
            )

    for col_idx in range(1, ws_in.max_column + 1):
        write_row = 3
        for read_row in range(3, ws_in.max_row + 1):
            cell = ws_in.cell(row=read_row, column=col_idx)
            val = str(cell.value).strip() if cell.value else ""

            if val == "=" or val == "":
                continue

            out_cell = ws_out.cell(row=write_row, column=col_idx)
            out_cell.value = val

            if cell.hyperlink:
                out_cell.hyperlink = cell.hyperlink.target
                out_cell.style = "Hyperlink"

                url = out_cell.hyperlink.target

                if len(url.split("/")) < 6:
                    technique_id = url.split("/")[-1]
                    out_cell.value = str(val) + "\n" + str(technique_id)


                else:
                    technique_id = url.split("/")[-2]
                    sub_technique_id = url.split("/")[-1]

                    out_cell.value = str(val) + "\n" + str(technique_id) + "\\" + str(sub_technique_id)

            if col_idx % 2 == 1:
                out_cell.font = Font(bold=True)
            elif re.search(r"\(\d+\)", val):
                out_cell.font = Font(bold=True)

            out_cell.border = thin_border
            out_cell.alignment = Alignment(
                wrap_text=True,
                horizontal='center',
                vertical='center'
            )

            write_row += 1

    for row in [1, 2]:

        # Replicate every odd column's value in row 1 to the next even column
        for col_idx in range(1, ws_in.max_column):
            if col_idx % 2 == 1:  # Check if column index is odd
                source_cell = ws_out.cell(row=row, column=col_idx)
                target_cell = ws_out.cell(row=row, column=col_idx + 1)

                # Copy value
                target_cell.value = source_cell.value

                # Copy hyperlink if present
                if source_cell.hyperlink:
                    target_cell.hyperlink = source_cell.hyperlink
                    target_cell.style = "Hyperlink"

                # Copy formatting
                # Copy font attributes individually
                source_font = source_cell.font
                target_cell.font = Font(
                    name=source_font.name,
                    size=source_font.size,
                    bold=source_font.bold,
                    italic=source_font.italic,
                    vertAlign=source_font.vertAlign,
                    underline=source_font.underline,
                    strike=source_font.strike,
                    color=source_font.color
                )

                # Copy border
                source_border = source_cell.border
                target_cell.border = Border(
                    left=source_border.left,
                    right=source_border.right,
                    top=source_border.top,
                    bottom=source_border.bottom
                )

                # Copy alignment
                source_alignment = source_cell.alignment
                target_cell.alignment = Alignment(
                    horizontal=source_alignment.horizontal,
                    vertical=source_alignment.vertical,
                    wrap_text=source_alignment.wrap_text
                )



    for col in ws_out.columns:
        col_letter = col[0].column_letter
        ws_out.column_dimensions[col_letter].width = 20


    wb_out.save(output_path)
    print(f"Cleaned MITRE matrix with bolded main techniques saved to:\n{output_path}")

    return